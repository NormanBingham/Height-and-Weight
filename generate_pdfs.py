import os
import sys
import argparse
from datetime import datetime
from pathlib import Path

# Check for required dependencies
try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is not installed.")
    print("Please install it using one of these commands:")
    print("  py -m pip install openpyxl fillpdf")
    print("  pip install openpyxl fillpdf")
    sys.exit(1)

try:
    from fillpdf import fillpdfs
except ImportError:
    print("Error: fillpdf is not installed.")
    print("Please install it using one of these commands:")
    print("  py -m pip install openpyxl fillpdf")
    print("  pip install openpyxl fillpdf")
    sys.exit(1)

def validate_file_paths(excel_file, pdf_5500, pdf_5501):
    """Validate that all input files exist and are accessible."""
    files_to_check = [
        (excel_file, "Excel file"),
        (pdf_5500, "PDF template 5500"),
        (pdf_5501, "PDF template 5501")
    ]
    
    for file_path, file_type in files_to_check:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"{file_type} not found: {file_path}")
        if not path.is_file():
            raise ValueError(f"{file_type} is not a file: {file_path}")
        if not os.access(file_path, os.R_OK):
            raise PermissionError(f"Cannot read {file_type}: {file_path}")

def safe_cell_value(cell, default=None):
    """Safely extract cell value with proper None handling."""
    return cell.value if cell.value is not None else default

def safe_float_conversion(value, default=0.0):
    """Safely convert value to float with error handling."""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default

def generate_pdfs(excel_file, pdf_5500, pdf_5501, output_dir, custom_date=None, debug=False):
    """Generate PDF forms from Excel data with comprehensive error handling."""
    try:
        # Validate input files
        validate_file_paths(excel_file, pdf_5500, pdf_5501)
        
        current_date = custom_date or datetime.today().strftime('%Y%m%d')

        HEIGHT_WEIGHT_PASS = (
            "The Soldier has met the Army's height and weight standards as outlined in AR 600-9. "
            "The Soldier's weight of {weight} pounds is within the maximum allowable weight of {max_weight} pounds."
            "The Soldier is encouraged to maintain current fitness and body composition levels to support mission readiness."
        )

        MET_STANDARD = (
            "The individual has met the Army's body fat standards as outlined in AR 600-9. "
            "The Soldier's body fat percentage was {body_fat_percentage}%, which is within the standard of {body_fat_standard}%."
            "The Soldier is in compliance with the Army Body Composition Program (ABCP) "
            "and requires no further action. Maintain focus on overall health and physical readiness."
        )

        DID_NOT_MEET_STANDARDS = (
            "The individual has exceeded the allowable body fat standards as outlined in AR 600-9. "
            "The Soldier's body fat percentage was {body_fat_percentage}%, which exceeds the standard of {body_fat_standard}%."
            "The Soldier is noncompliant with the Army Body Composition Program (ABCP) and must be enrolled in the program. "
            "Soldier will adhere to the requirements outlined in AR 600-9 to achieve compliance. "
        )

        ACFT_FAIL_HEIGHT_WEIGHT_PASS = (
            "The Soldier exceeded the Army's height and weight standards outlined in AR 600-9; however, "
            "the Soldier achieved a total score of 540 or above on the Army Combat Fitness Test (ACFT), "
            "with a minimum of 80 points in each event. As such, the Soldier is exempt from being flagged or enrolled in the Army Body Composition Program (ABCP)."
        )

        # Load workbook with error handling
        try:
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = workbook.active
        except Exception as e:
            raise ValueError(f"Failed to load Excel file: {e}")

        custom_mapping = {
            0: "NAME",
            1: "RANK",
            3: "AGE",
            4: "HEIGHT",
            5: "WEIGHT",  
            9: "FIRST",
            10: "SECOND",
            11: "THIRD",
            12: ["AVERAGE", "AVERAGE_1"],  
            13: "BODY FAT PERCENTAGE",
            17: "PREPARED BY",
            18: "PREP_BY_RANK",
            19: "APPROVED BY SUPERVISOR",
            20: "APPR_BY_RANK",
        }

        # Extract default values safely
        default_prepared_by = safe_cell_value(sheet.cell(row=2, column=1), "")
        default_prep_by_rank = safe_cell_value(sheet.cell(row=2, column=2), "")
        preparer_initials = safe_cell_value(sheet.cell(row=2, column=3), "")
        default_approved_by_supervisor = safe_cell_value(sheet.cell(row=4, column=1), "")
        default_appr_by_rank = safe_cell_value(sheet.cell(row=4, column=2), "")

        # Create output directory safely
        try:
            Path(output_dir).mkdir(parents=True, exist_ok=True)
        except Exception as e:
            raise PermissionError(f"Cannot create output directory: {e}")

        print("Starting PDF generation...")
        processed_count = 0
        error_count = 0

        for row in sheet.iter_rows(min_row=6, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            if not safe_cell_value(row[0]):
                print(f"Stopping at row {row[0].row} as the first column is empty.")
                break

            try:
                # Check for row-specific overrides (columns R=17, S=18, T=19, U=20)
                row_prepared_by = safe_cell_value(row[17], default_prepared_by)
                row_prep_by_rank = safe_cell_value(row[18], default_prep_by_rank)
                row_approved_by_supervisor = safe_cell_value(row[19], default_approved_by_supervisor)
                row_appr_by_rank = safe_cell_value(row[20], default_appr_by_rank)

                gender = safe_cell_value(row[2], "").strip().upper()
                name = safe_cell_value(row[0], f"Unknown_{row[0].row}")
                
                if gender == "M":
                    pdf_template = pdf_5500
                    output_filename = str(Path(output_dir) / f"{name}_5500.pdf")
                elif gender == "F":
                    pdf_template = pdf_5501
                    output_filename = str(Path(output_dir) / f"{name}_5501.pdf")
                else:
                    print(f"Warning: Unknown gender '{gender}' at row {row[0].row}. Skipping.")
                    error_count += 1
                    continue

                data_dict = {
                    "DATE1": current_date,
                    "DATE2": current_date,
                }

                # Use row-specific values (with fallback to defaults)
                data_dict["PREPARED BY"] = row_prepared_by
                data_dict["PREP_BY_RANK"] = row_prep_by_rank
                data_dict["APPROVED BY SUPERVISOR"] = row_approved_by_supervisor
                data_dict["APPR_BY_RANK"] = row_appr_by_rank

                stop_reading_row = False
                height_weight_pass = False
                exemption = False

                for col_idx, cell in enumerate(row):
                    if col_idx == 8 and str(safe_cell_value(cell, "")).strip() == "Pass":
                        if debug:
                            print(f"Column 9 (index 8) is 'Pass' — stopping further reading for this row at column {col_idx}.")
                        stop_reading_row = True
                        height_weight_pass = True
                        break
                    # Check if column 7 ("Yes") and Fail Tape to set special exemption
                    elif safe_cell_value(row[6]) == "Yes" and safe_cell_value(row[8]) == "Needs Tape":
                        if debug:
                            print(f"Column 7 (index 6) is 'Yes' and Column 9 (index 8) is Needs Tape — stopping further reading for this row at column {col_idx}.")
                        stop_reading_row = True
                        exemption = True
                        remarks = ACFT_FAIL_HEIGHT_WEIGHT_PASS
                        data_dict["REMARKS"] = remarks
                        if debug:
                            print(f"Writing ACFT_FAIL_HEIGHT_WEIGHT_PASS remark: {remarks}")
                    elif safe_cell_value(row[6]) == "Yes" and safe_cell_value(row[8]) == "Needs Tape" and safe_cell_value(row[16]) == "Fail Tape":
                        if debug:
                            print(f"Column 7 (index 6) is 'Yes' and Column 9 (index 8) is Needs Tape — stopping further reading for this row at column {col_idx}.")
                        stop_reading_row = True
                        exemption = True
                        remarks = ACFT_FAIL_HEIGHT_WEIGHT_PASS
                        data_dict["REMARKS"] = remarks
                        if debug:
                            print(f"Writing ACFT_FAIL_HEIGHT_WEIGHT_PASS remark: {remarks}")

                    if safe_cell_value(cell) is not None:
                        form_field_key = custom_mapping.get(col_idx)
                        if form_field_key:
                            if isinstance(form_field_key, list):
                                for key in form_field_key:
                                    data_dict[key] = cell.value
                                    if debug:
                                        print(f"Writing '{cell.value}' to form field '{key}'")
                            else:
                                data_dict[form_field_key] = cell.value
                                if debug:
                                    print(f"Writing '{cell.value}' to form field '{form_field_key}'")

                if exemption and safe_cell_value(row[8]) == "Needs Tape":
                    data_dict["Preparer's Initials"] = preparer_initials
                    if debug:
                        print(f"Added 'Preparer's Initials': {preparer_initials} because column 6 is 'Yes'")

                if not stop_reading_row:
                    body_fat_percentage = safe_float_conversion(row[13].value)
                    body_fat_standard = safe_float_conversion(row[15].value)

                    data_dict["BODY FAT PERCENTAGE"] = int(body_fat_percentage * 100)
                    data_dict["BODY FAT STANDARD"] = int(body_fat_standard * 100)
                else:
                    if debug:
                        print(f"Skipping body fat and standard values for row {row[0].row} because column 9 is 'Pass'.")

                if height_weight_pass:
                    weight = safe_cell_value(row[5])
                    max_weight = safe_cell_value(row[7])
                    if weight is not None and max_weight is not None:
                        remarks = HEIGHT_WEIGHT_PASS.format(weight=weight, max_weight=max_weight)
                        data_dict["REMARKS"] = remarks
                        if debug:
                            print(f"Writing HEIGHT_WEIGHT_PASS remark: {remarks}")
                    else:
                        print(f"Warning: max weight (column 8) missing for {name}, cannot fill remarks properly.")
                elif safe_cell_value(row[8]) == "Needs Tape" and safe_cell_value(row[16]) == "Fail Tape" and not exemption:
                    body_fat_percentage = safe_float_conversion(row[13].value)
                    body_fat_standard = safe_float_conversion(row[15].value)
                    data_dict["BODY FAT PERCENTAGE"] = int(body_fat_percentage * 100)
                    data_dict["BODY FAT STANDARD"] = int(body_fat_standard * 100)
                    remarks = DID_NOT_MEET_STANDARDS.format(
                        body_fat_percentage=data_dict["BODY FAT PERCENTAGE"],
                        body_fat_standard=data_dict["BODY FAT STANDARD"],
                    )
                    data_dict["REMARKS"] = remarks
                    if debug:
                        print(f"Writing DID_NOT_MEET_STANDARDS remark: {remarks}")
                elif safe_cell_value(row[8]) == "Needs Tape" and safe_cell_value(row[16]) == "Pass" and not exemption:
                    body_fat_percentage = safe_float_conversion(row[13].value)
                    body_fat_standard = safe_float_conversion(row[15].value)
                    data_dict["BODY FAT PERCENTAGE"] = int(body_fat_percentage * 100)
                    data_dict["BODY FAT STANDARD"] = int(body_fat_standard * 100)
                    remarks = MET_STANDARD.format(
                        body_fat_percentage=data_dict["BODY FAT PERCENTAGE"],
                        body_fat_standard=data_dict["BODY FAT STANDARD"],
                    )
                    data_dict["REMARKS"] = remarks
                    if debug:
                        print(f"Writing MET_STANDARD remark: {remarks}")

                if safe_cell_value(row[8]) == "Needs Tape" and safe_cell_value(row[5]):
                    data_dict["WEIGHT_1"] = row[5].value
                    if debug:
                        print(f"Added 'WEIGHT_1': {row[5].value}")

                if safe_cell_value(row[8]) == "Pass":
                    data_dict["IN COMPLIANCE"] = "Yes"
                    if debug:
                        print("Added check for compliance: Yes")
                elif safe_cell_value(row[16]) == "Pass":
                    data_dict["IN COMPLIANCE"] = "Yes"
                    if debug:
                        print("Added check for compliance: Yes")
                elif safe_cell_value(row[6]) == "Yes" and safe_cell_value(row[16]) == "Fail Tape":
                    data_dict["IN COMPLIANCE"] = "Yes"
                    if debug:
                        print("Added check for compliance: Yes")
                elif safe_cell_value(row[16]) == "Fail Tape" and not exemption:
                    data_dict["NOT IN COMPLIANCE"] = "Yes"
                    if debug:
                        print("Added check for non-compliance: Yes")

                # Generate PDF with error handling
                try:
                    fillpdfs.write_fillable_pdf(
                        input_pdf_path=pdf_template,
                        output_pdf_path=output_filename,
                        data_dict=data_dict,
                    )
                    print(f"Generated: {output_filename}")
                    processed_count += 1
                except Exception as e:
                    print(f"Error generating PDF for {name}: {e}")
                    error_count += 1
                    continue
                    
            except Exception as e:
                print(f"Error processing row {row[0].row}: {e}")
                error_count += 1
                continue

        print(f"PDF generation complete! Processed: {processed_count}, Errors: {error_count}")
        
    except Exception as e:
        print(f"Fatal error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate individual PDF forms from Excel data using PDF templates.",
        epilog="""Examples:
  Basic usage:
    py generate_pdfs.py --excel "data.xlsx" --pdf_5500 "male_form.pdf" --pdf_5501 "female_form.pdf" --output "output_folder"

  With custom date:
    py generate_pdfs.py --excel "data.xlsx" --pdf_5500 "male_form.pdf" --pdf_5501 "female_form.pdf" --output "output_folder" --date "2024-01-15"

  Full Windows path example:
    py generate_pdfs.py --excel "C:\\Users\\user\\Desktop\\data.xlsx" --pdf_5500 "C:\\Users\\user\\Desktop\\Forms\\Male_Form.pdf" --pdf_5501 "C:\\Users\\user\\Desktop\\Forms\\Female_Form.pdf" --output "C:\\Users\\user\\Desktop\\Output"

Required Python packages:
  Install with: py -m pip install openpyxl fillpdf
  Or: pip install openpyxl fillpdf

Template files and Excel format:
  Download the required Excel template and PDF forms from:
  https://github.com/NormanBingham/Height-and-Weight
""",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--excel", required=True, help="Path to the Excel file (use template from GitHub repo)")
    parser.add_argument("--pdf_5500", required=True, help="Path to the PDF template for Males (Form 5500 from GitHub repo)")
    parser.add_argument("--pdf_5501", required=True, help="Path to the PDF template for Females (Form 5501 from GitHub repo)")
    parser.add_argument("--output", required=True, help="Directory to save generated PDFs (created automatically if needed)")
    parser.add_argument("--date", help="Custom date for forms (format: YYYY-MM-DD, defaults to today)")
    parser.add_argument("--debug", action="store_true", help="Enable verbose debug output")

    args = parser.parse_args()

    custom_date = None
    if args.date:
        try:
            parsed_date = datetime.strptime(args.date, "%Y-%m-%d")
            custom_date = parsed_date.strftime("%Y%m%d")
        except ValueError as e:
            print(f"Error: Invalid date format '{args.date}'. Use YYYY-MM-DD format.")
            sys.exit(1)

    try:
        generate_pdfs(
            excel_file=args.excel,
            pdf_5500=args.pdf_5500,
            pdf_5501=args.pdf_5501,
            output_dir=args.output,
            custom_date=custom_date,
            debug=args.debug,
        )
    except KeyboardInterrupt:
        print("\nOperation cancelled by user.")
        sys.exit(1)
    except Exception as e:
        print(f"Application error: {e}")
        sys.exit(1)